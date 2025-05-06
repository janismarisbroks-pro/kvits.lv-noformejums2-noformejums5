import os
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from django.core.management.base import BaseCommand
from django.conf import settings
from kvitsapp.models import Product
from decimal import Decimal

class Command(BaseCommand):
    help = 'Import products from Excel and extract images'

    def _preprocess_image_mapping(self, sheet, image_loader):
        """Build a mapping from row numbers to image cell addresses"""
        row_to_image = {}
        
        # First, find all images in the sheet
        image_cells = {}
        for row in range(1, sheet.max_row + 1):
            for cell in sheet[row]:
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    cell_address = cell.coordinate
                    if image_loader.image_in(cell_address):
                        image_cells[cell_address] = (row, cell.column)
        
        # Then, map rows to images including merged cell ranges
        for cell_address, (img_row, img_col) in image_cells.items():
            # Check if this cell is part of a merged range
            for merged_range in sheet.merged_cells.ranges:
                start_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                if start_cell.coordinate == cell_address:
                    # Map all rows in the merged range to this image
                    for r in range(merged_range.min_row, merged_range.max_row + 1):
                        row_to_image[r] = cell_address
                    break
            else:
                # This is a regular cell with an image
                row_to_image[img_row] = cell_address
                
                # Also check if this image should apply to subsequent rows without images
                # (visual association rather than merged cells)
                for r in range(img_row + 1, min(img_row + 5, sheet.max_row + 1)):
                    # Don't override rows that already have images assigned
                    if r not in row_to_image:
                        # Check if this row has product data but no image
                        if sheet.cell(row=r, column=1).value or sheet.cell(row=r, column=2).value:
                            # Look ahead for next image
                            has_own_image = False
                            for check_row in range(r, min(r + 3, sheet.max_row + 1)):
                                if check_row in row_to_image and check_row != img_row:
                                    has_own_image = True
                                    break
                            
                            if not has_own_image:
                                row_to_image[r] = cell_address
        
        return row_to_image

    def _find_image_for_row(self, sheet, image_loader, row_index, row_to_image_map):
        """Find image for a specific row using the pre-built mapping"""
        # First check the mapping
        if row_index in row_to_image_map:
            cell_address = row_to_image_map[row_index]
            
            # Determine if this is from a merged cell
            is_merged = False
            for merged_range in sheet.merged_cells.ranges:
                if (merged_range.min_row <= row_index <= merged_range.max_row and
                    sheet.cell(row=merged_range.min_row, column=merged_range.min_col).coordinate == cell_address):
                    is_merged = True
                    break
                    
            return cell_address, is_merged
        
        # Fallback: look for images in surrounding rows if not in mapping
        for check_row in range(row_index, max(1, row_index - 10), -1):
            for cell in sheet[check_row]:
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    cell_address = cell.coordinate
                    if image_loader.image_in(cell_address):
                        return cell_address, check_row != row_index
        
        return None, False

    def handle(self, *args, **kwargs):
        filepath = os.path.join('katalogs_2024.xlsx')
        output_folder = os.path.join(settings.BASE_DIR, 'kvitsapp', 'static', 'images')
        os.makedirs(output_folder, exist_ok=True)

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = wb.active
            image_loader = SheetImageLoader(sheet)
            
            # Pre-process to build row to image mapping
            self.stdout.write(self.style.SUCCESS('Pre-processing sheet to map rows to images...'))
            row_to_image_map = self._preprocess_image_mapping(sheet, image_loader)
            self.stdout.write(self.style.SUCCESS(f'Found {len(row_to_image_map)} rows with associated images'))
            
            # Cache for images we've already processed
            processed_images = {}
            # Cache for mapping cell addresses to file paths
            cell_to_filename = {}

            for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                try:
                    ean13 = row[0].value
                    pasutijuma_kods = row[1].value
                    
                    # Skip rows without product code
                    if not pasutijuma_kods:
                        continue
                        
                    apraksts = row[3].value
                    cena_raw_value = row[6].value

                    # Process price
                    if cena_raw_value is not None:
                        try:
                            cena = Decimal(str(cena_raw_value).replace(',', '.'))
                        except (ValueError, TypeError) as e:
                            self.stdout.write(
                                self.style.ERROR(f'Error processing price in row {row_index}: {e}')
                            )
                            continue
                    else:
                        self.stdout.write(
                            self.style.WARNING(f'Price missing in row {row_index} for {pasutijuma_kods}. Skipping.')
                        )
                        continue

                    # Find image for this row using our mapping
                    cell_address, is_merged = self._find_image_for_row(
                        sheet, image_loader, row_index, row_to_image_map
                    )

                    # Handle image
                    if cell_address:
                        # Check if we already processed this image cell
                        if cell_address in cell_to_filename:
                            attels_filename = cell_to_filename[cell_address]
                            self.stdout.write(f'Using previously saved image for {pasutijuma_kods} from cell {cell_address}')
                        else:
                            # Extract and save new image
                            image = image_loader.get(cell_address)
                            filename = f"product_{pasutijuma_kods}.png"
                            image_path = os.path.join(output_folder, filename)
                            image.save(image_path)
                            attels_filename = f"images/{filename}"
                            cell_to_filename[cell_address] = attels_filename
                            self.stdout.write(
                                self.style.SUCCESS(
                                    f'Extracted and saved image from {"merged " if is_merged else ""}cell {cell_address} for {pasutijuma_kods}'
                                )
                            )
                    else:
                        attels_filename = "images/no-image.png"
                        self.stdout.write(
                            self.style.WARNING(f'No image found for {pasutijuma_kods} in row {row_index}')
                        )

                    # Create or update product
                    Product.objects.update_or_create(
                        pasutijuma_kods=pasutijuma_kods,
                        defaults={
                            'ean13': ean13 or '',
                            'attels': attels_filename,
                            'apraksts': apraksts or '',
                            'cena': cena
                        }
                    )
                    self.stdout.write(f'Imported product {pasutijuma_kods}')

                except Exception as e:
                    self.stdout.write(self.style.ERROR(f'Error processing row {row_index}: {e}'))

            self.stdout.write(self.style.SUCCESS(f'Products imported successfully! Processed {len(cell_to_filename)} unique images.'))

        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f'Error: The file "{filepath}" was not found.'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'An unexpected error occurred: {e}'))


#docker-compose exec web python manage.py import_products
#docker-compose exec web python manage.py flush

#lai apskatītu datubāzi, var izmantot šādu komandu:
#docker-compose exec web bash
    # apt-get update
    # apt-get install -y postgresql-client
#psql -h db -U postgres
# SELECT * FROM kvitsapp_product;
#\q